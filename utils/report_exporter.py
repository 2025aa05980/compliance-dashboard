"""
Report export utility. Supports CSV and multi-sheet Excel.
Extend with openpyxl conditional formatting for production.
"""
import pandas as pd
import io
import os
from datetime import datetime

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports_output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

STATUS_COLORS_XL = {
    "Compliant":           "FF90EE90",
    "Non-Compliant":       "FFFFB3B3",
    "Partially Compliant": "FFFFFF99",
    "Exception Approved":  "FFD8BFD8",
    "Not Assessed":        "FFD3D3D3",
}

RISK_COLORS_XL = {
    "Critical": "FFFFB3B3",
    "High":     "FFFFFF99",
    "Medium":   "FFADD8E6",
    "Low":      "FF90EE90",
}


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8")


def df_to_excel_bytes(dfs: dict[str, pd.DataFrame], title: str = "IAM Compliance Report") -> bytes:
    """
    dfs: { sheet_name: dataframe }
    Returns bytes of an xlsx workbook with one sheet per dataframe,
    conditional colour-coding on Compliance_Status and Risk_Rating columns.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="FF1F3A5F")
    header_font = Font(color="FFFFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="FFD3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, df in dfs.items():
        # Truncate sheet names (Excel limit = 31 chars)
        ws = wb.create_sheet(title=sheet_name[:31])
        # Write header
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, (col_name, val) in enumerate(zip(df.columns, row), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_safe_val(val))
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                # Apply status colour
                if col_name == "Compliance_Status" and str(val) in STATUS_COLORS_XL:
                    cell.fill = PatternFill("solid", fgColor=STATUS_COLORS_XL[str(val)])
                elif col_name == "Risk_Rating" and str(val) in RISK_COLORS_XL:
                    cell.fill = PatternFill("solid", fgColor=RISK_COLORS_XL[str(val)])
            # Alternating row background
            if row_idx % 2 == 0:
                for col_idx in range(1, len(df.columns) + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    if c.fill.fgColor.rgb == "00000000":
                        c.fill = PatternFill("solid", fgColor="FFF8F8F8")

        # Auto-size columns (cap at 40)
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe_val(val):
    if pd.isna(val) if not isinstance(val, (str, bool)) else False:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    return val


def generate_summary_excel(all_summaries: dict) -> bytes:
    """Single workbook with all domain summaries."""
    dfs = {}
    for domain, summary in all_summaries.items():
        if isinstance(summary, pd.DataFrame):
            dfs[domain[:31]] = summary
    return df_to_excel_bytes(dfs)
